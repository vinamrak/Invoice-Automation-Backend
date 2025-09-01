import io
import os
import platform
import shutil
import subprocess
import tempfile
from datetime import datetime
from calendar import monthrange

import fitz  # PyMuPDF
import pytz
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from fastapi import FastAPI, BackgroundTasks
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.date import DateTrigger

from apscheduler.triggers.cron import CronTrigger


# ---------- Config ----------

INVOICE_CONFIGS = {
    "Reliance Footwear": {
        "excel": "FootwearInvoice.xlsx",
        "signature": "VinamraRealtorsStampAndVineetSignature.png",
        "pdf_name": "Reliance_Footwear_Invoice.pdf",
        "recipient": "sanjiban1.das@ril.com",
        "cc": "to-nirbhik.jana@ril.com,vinamrakhoria@gmail.com",
        "subject": "Footwear Invoice for ",
        "invoice_code": "Foot",
        "signature_position": (600, 380, 120, 120),
    },
    "Reliance Jewels": {
        "excel": "JewelsInvoice.xlsx",
        "signature": "VinamraRealtorsStampAndVineetSignature.png",
        "pdf_name": "Reliance_Jewels_Invoice.pdf",
        "recipient": "bhaskar.banik@ril.com",
        "cc": "dinesh.deshmane@ril.com,vinamrakhoria@gmail.com",
        "subject": "Jewels Invoice for ",
        "invoice_code": "Jewel",
        "signature_position": (600, 380, 120, 120),
    },
}

## Default signature placement (points) if not specified in config
SIG_X, SIG_Y, SIG_W, SIG_H = 600, 380, 120, 120
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "*").split(",")

app = FastAPI(title="Invoice Service")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in ALLOWED_ORIGINS],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def update_excel_inplace(path: str, invoice_code: str = "BIG") -> None:
    wb = load_workbook(path)
    ws = wb.active

    # today = datetime.today()
    today = datetime(2026, 1, 1) 
    current_month = today.month
    current_year = today.year
    year_two_digits = str(current_year)[-2:]

    if current_month >= 4:
        fy_month_number = current_month - 3
        fy_start_year = current_year
        fy_end_year = current_year + 1
    else:
        fy_month_number = current_month + 9
        fy_start_year = current_year - 1
        fy_end_year = current_year

    fy_start_short = str(fy_start_year)[-2:]
    fy_end_short = str(fy_end_year)[-2:]

    ws["A9"] = f"Invoice Number- {fy_month_number}/{invoice_code}/{fy_start_short}-{fy_end_short}"
    ws["J9"] = today.replace(day=1).strftime("%d/%m/%Y")
    month_shorthand = today.strftime("%b")
    ws["A22"] = f"Rent for the month of {month_shorthand},{year_two_digits}"

    last_day = monthrange(current_year, current_month)[1]
    first_day_str = f"01/{current_month:02d}/{current_year}"
    last_day_str = f"{last_day:02d}/{current_month:02d}/{current_year}"
    ws["A23"] = f"({first_day_str} - {last_day_str})"
    ws["D23"] = last_day  # Set D23 to number of days in the month
    wb.save(path)

def convert_xlsx_to_pdf(input_xlsx: str, out_dir: str) -> str:
    # Ensure LibreOffice path (Dockerfile puts 'soffice' on PATH)
    if platform.system() == "Darwin":
        soffice = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    else:
        soffice = "soffice"

    try:
        subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", input_xlsx, "--outdir", out_dir],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
    except subprocess.CalledProcessError as e:
        raise HTTPException(status_code=500, detail=f"LibreOffice conversion failed: {e.stderr.decode(errors='ignore')}")

    pdf_path = os.path.join(
        out_dir, os.path.splitext(os.path.basename(input_xlsx))[0] + ".pdf"
    )
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=500, detail="PDF not generated.")
    return pdf_path

def add_signature_bytes(pdf_path: str, signature_path: str, signature_position=None) -> bytes:
    if not os.path.exists(signature_path):
        raise HTTPException(status_code=500, detail="Signature image not found on server.")

    if signature_position is None:
        x, y, w, h = SIG_X, SIG_Y, SIG_W, SIG_H
    else:
        x, y, w, h = signature_position

    doc = fitz.open(pdf_path)
    try:
        page = doc[0]
        rect = fitz.Rect(x, y, x + w, y + h)
        page.insert_image(rect, filename=signature_path)
        return doc.tobytes()  # in-memory bytes
    finally:
        doc.close()


    # """
    # Generates the invoice PDF (with signature) for a given config and returns it as bytes.
    # Raises HTTPException if template or signature is missing.
    # """
    # excel_template = config["excel"]
    # signature_image = config["signature"]
    # invoice_code = config.get("invoice_code", "BIG")
    # signature_position = config.get("signature_position", (SIG_X, SIG_Y, SIG_W, SIG_H))
    # if not os.path.exists(excel_template):
    #     raise HTTPException(status_code=500, detail=f"{excel_template} not found on server.")
    # if not os.path.exists(signature_image):
    #     raise HTTPException(status_code=500, detail=f"{signature_image} not found on server.")

    # with tempfile.TemporaryDirectory() as tmpdir:
    #     working_xlsx = os.path.join(tmpdir, os.path.basename(excel_template))
    #     shutil.copy(excel_template, working_xlsx)
    #     update_excel_inplace(working_xlsx, invoice_code=invoice_code)
    #     pdf_path = convert_xlsx_to_pdf(working_xlsx, tmpdir)
    #     final_pdf_bytes = add_signature_bytes(pdf_path, signature_image, signature_position=signature_position)
    # return final_pdf_bytes



# --- Move generate_invoice_pdf_bytes definition above its first use ---

def generate_invoice_pdf_bytes(config: dict) -> bytes:
    """
    Generates the invoice PDF (with signature) for a given config and returns it as bytes.
    Raises HTTPException if template or signature is missing.
    """
    excel_template = config["excel"]
    signature_image = config["signature"]
    invoice_code = config.get("invoice_code", "BIG")
    signature_position = config.get("signature_position", (SIG_X, SIG_Y, SIG_W, SIG_H))
    if not os.path.exists(excel_template):
        raise HTTPException(status_code=500, detail=f"{excel_template} not found on server.")
    if not os.path.exists(signature_image):
        raise HTTPException(status_code=500, detail=f"{signature_image} not found on server.")

    with tempfile.TemporaryDirectory() as tmpdir:
        working_xlsx = os.path.join(tmpdir, os.path.basename(excel_template))
        shutil.copy(excel_template, working_xlsx)
        update_excel_inplace(working_xlsx, invoice_code=invoice_code)
        pdf_path = convert_xlsx_to_pdf(working_xlsx, tmpdir)
        final_pdf_bytes = add_signature_bytes(pdf_path, signature_image, signature_position=signature_position)
    return final_pdf_bytes


@app.get("/download-all-invoices")
def download_all_invoices():
    # Returns both PDFs as a zip
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for key, config in INVOICE_CONFIGS.items():
            pdf_bytes = generate_invoice_pdf_bytes(config)
            zf.writestr(config["pdf_name"], pdf_bytes)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="All_Invoices.zip"'}
    return StreamingResponse(buf, media_type="application/zip", headers=headers)

@app.head("/ping")
def ping():
    return {"Status":"Server up and running :)"}


SMTP_SERVER = "smtp.mail.yahoo.com"
SMTP_PORT = 587
SENDER_EMAIL = "vineetkhoria@yahoo.com"
SENDER_PASSWORD = "kecs ftif zavs xzqu"  # Use App Password for Gmail, not real password


def send_email_with_pdf(to_email: str, subject: str, body: str, pdf_path: str, pdf_name: str, cc_email: str = None):
    # Create the email
    msg = MIMEMultipart()
    msg["From"] = SENDER_EMAIL
    msg["To"] = to_email
    msg["Subject"] = subject
    if cc_email:
        msg["Cc"] = cc_email

    # Add body
    msg.attach(MIMEText(body, "plain"))

    # Attach PDF with configurable filename
    with open(pdf_path, "rb") as f:
        pdf = MIMEApplication(f.read(), _subtype="pdf")
        pdf.add_header("Content-Disposition", "attachment", filename=pdf_name)
        msg.attach(pdf)

    # Send email
    recipients = [to_email]
    if cc_email:
        recipients.append(cc_email)
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()  # Secure connection
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, recipients, msg.as_string())



def trigger_email(config: dict, background_tasks: BackgroundTasks):
    # Generate PDF bytes and save to temp file
    final_pdf_bytes = generate_invoice_pdf_bytes(config)
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
        tmp_pdf.write(final_pdf_bytes)
        tmp_pdf_path = tmp_pdf.name
    # Compose subject with month and year
    # today = datetime.today()
    today = datetime(2026, 1, 1) 
    month_year = today.strftime("%B %Y")
    subject = f"{config['subject']} {month_year}"
    body = f"Hello, please find attached your invoice for {config['pdf_name'].split('_')[0]}."
    background_tasks.add_task(
        send_email_with_pdf,
        config["recipient"],
        subject,
        body,
        tmp_pdf_path,
        config["pdf_name"],
        config.get("cc")
    )
    return {"message": f"Email is being sent to {config['recipient']}!"}


def send_all_invoices():
    print("Sending all invoice emails")
    # today = datetime.today()
    today = datetime(2026, 1, 1) 
    month_year = today.strftime("%B %Y")
    for key, config in INVOICE_CONFIGS.items():
        final_pdf_bytes = generate_invoice_pdf_bytes(config)
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
            tmp_pdf.write(final_pdf_bytes)
            tmp_pdf_path = tmp_pdf.name
        try:
            subject = f"{config['subject']} {month_year}"
            send_email_with_pdf(
                config["recipient"],
                subject,
                f"Hello, please find attached your invoice for {key}.",
                tmp_pdf_path,
                config["pdf_name"],
                config.get("cc")
            )
        finally:
            if os.path.exists(tmp_pdf_path):
                os.remove(tmp_pdf_path)

# Initialize scheduler
scheduler = BackgroundScheduler(timezone=pytz.timezone("Asia/Kolkata"))

# Schedule to run at 8:00AM on the first day of every month
scheduler.add_job(
    send_all_invoices,
    trigger=CronTrigger(day=1, hour=10, minute=35, second=0, timezone=pytz.timezone("Asia/Kolkata")),
    name="Send invoices monthly at 8:00AM IST on the 1st"
)
scheduler.start()
